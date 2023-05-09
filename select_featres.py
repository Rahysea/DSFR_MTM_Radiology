'''
        Filename:Classifier_Train_Atsea.py
        User: Atsea
        Create: 2022/8/25
        Modifier: Atsea
'''
from sklearn import feature_selection
from sklearn.metrics import mean_squared_error
from sklearn.linear_model import Lasso
from openpyxl import Workbook
from openpyxl import load_workbook
import csv
import argparse
import h5py
import os
import gc
import warnings
import pandas as pd
import random
import numpy as np
warnings.filterwarnings('ignore')
sep = os.sep
filesep = sep

class Classfier():
    def __init__(self):
        super().__init__()

    def dict2np(self, data, dsfr = False):
        """
        :param data: A dictionary containing all patient radiomics, deep learning features, and labels
        :param dsfr: A boolean value indicating whether there are deep learning features present
        :return:
        """
        all_key = data.keys()
        infomation_list = []
        radiofeature_list = []
        dsfrfeature_list = []
        for info in all_key:
            infomation = info.split('_')
            phase_parameter = infomation[0] + '_'+ infomation[1]
            radiofeature = data[info]['radio_feature']
            id = infomation[2]
            label = str(data[info]['label'])
            infomation_list.append([phase_parameter, id, label])
            radiofeature_list.append(list(radiofeature))
            if dsfr:
                dsfrfeature = data[info]['dsfr_feature']
                dsfrfeature_list.append(list(dsfrfeature))

        radio_feature = np.array(radiofeature_list)
        if dsfr:
            dsfr_feature = np.array(dsfrfeature_list)
        else:
            dsfr_feature = []
        infomation = np.array(infomation_list)
        return infomation, dsfr_feature, radio_feature

    def lasso(self, Data, label, alpha):
        model_lasso = Lasso(alpha=alpha, max_iter=2).fit(Data, label)
        coef = model_lasso.coef_
        intercept = model_lasso.intercept_
        # mse
        pred = model_lasso.predict(Data)
        mse = mean_squared_error(label, pred)

        return coef, intercept, mse

    def feature_choose(self, Data, Data_valid, Data_out, label, pers_chi=5, test_outside=True):
        function_choose = feature_selection.mutual_info_classif2
        fs = feature_selection.SelectKBest(function_choose, k=pers_chi)
        feature = fs.fit_transform(Data, label)
        random_state_save = str(fs.pvalues_)
        save_index = fs.get_support(indices=True)
        if test_outside:
            feature_valid = fs.transform(Data_valid)
            feature_out_sy = fs.transform(Data_out)
        # print('final feature num: ', feature.shape[1])
        return feature, feature_valid, feature_out_sy, random_state_save, save_index

    def choose_feature(self, feature, label, feature_out=None, feature_valid= None, alpha=0.0008):
        # lasso features select
        # External data is not available by default
        test_outside = False
        if feature_out is not None:
            test_outside = True
        alpha = alpha
        print('select features first, alpha: ', alpha)

        coel, intercept, mse = self.lasso(feature, label, alpha)
        fl = 0
        for i in range(len(coel)):
            if coel[i] != 0:
                fl = fl + 1
        print('lasso significant feature num:', fl)
        feature = feature[:, coel != 0]

        if test_outside:
            feature_out = feature_out[:, coel != 0]
            feature_validation = feature_valid[:, coel != 0]
        return mse, feature, feature_validation, feature_out, fl, coel, intercept

    def writesimple(self, path, data, index, column, sheetname='Sheet'):
        if os.path.exists(path):
            bg = load_workbook(path)
            sheets = bg.sheetnames
            if sheetname in sheets:
                sheet = bg[sheetname]
                sheet.cell(index, column, data)
                bg.save(path)
                bg.close()
            else:
                sheet = bg.create_sheet(sheetname)
                sheet.cell(index, column, data)
                bg.save(path)
                bg.close()
        else:
            bg = Workbook()
            bg1 = bg['Sheet']
            bg1.title = sheetname
            sheet = bg[sheetname]
            sheet.cell(index, column, data)
            bg.save(path)
            bg.close()

    def run(self, args):
        random.seed(417)
        Inner_PATIENTS_Label = {}
        Out_PATIENTS_Label_center2 = {}
        os.makedirs(args.savepath, exist_ok=True)
        # center 1
        in_p_num = 0 #  Count of '1's in training dataset
        # center 2 and center 3
        out_p_num_sy = 0  # Count of '1's in testing dataset
        # Reading the content of the data information file
        alldata_label = pd.read_csv(args.split_train_test)  # read dataset information
        # only use the 'use' data
        in_info = alldata_label.loc[alldata_label['dataset'] != 0]  # train information
        test_info = alldata_label.loc[alldata_label['dataset'] == 0]  # test information'

        in_patient_id = in_info['ID'].values.tolist()
        test_patient_id = test_info['ID'].values.tolist()

        in_patient_label = in_info['MTM'].values.tolist()
        in_type = in_info['dataset'].values.tolist()
        test_patient_label = test_info['MTM'].values.tolist()

        for i in range(len(in_patient_id)):
            Inner_PATIENTS_Label[in_patient_id[i]] = in_patient_label[i] # train ID and labels
            if in_patient_label[i] == 1:
                in_p_num = in_p_num + 1

        for i in range(len(test_patient_id)):
            Out_PATIENTS_Label_center2[test_patient_id[i]] = test_patient_label[i] # test ID and labels
            if test_patient_label[i] == 1:
                out_p_num_sy = out_p_num_sy + 1

        # Loading the internal data for center 1
        In_data = {}
        for pth_id in Inner_PATIENTS_Label.keys():
            if 'radio' in args.data_name:
                radio_parameter_name_path = os.path.join(args.radio_feature_path, args.mapname)
                H5_file_radio = h5py.File(os.path.join(radio_parameter_name_path, str(pth_id) + '.h5'), 'r')
                radio_feature = H5_file_radio['f_values'][:][0]
                H5_file_radio.close()

                key = args.mapname + '_' + str(pth_id)
                In_data[key] = {}
                In_data[key]['radio_feature'] = radio_feature
                In_data[key]['label'] = Inner_PATIENTS_Label[str(pth_id)]

                if 'open-mtm-dsfr' in args.data_name:
                    dsfr_parameter_name_path = os.path.join(args.dsfr_feature_path, args.mapname)
                    H5_file_dsfr = h5py.File(os.path.join(dsfr_parameter_name_path, str(pth_id) + '.h5'), 'r')
                    dsfr_feature = H5_file_dsfr['f_values'][:]
                    H5_file_dsfr.close()
                    In_data[key]['dsfr_feature'] = dsfr_feature

            if len(args.data_name) == 0:
                raise Exception("No data!")

        # Loading the test data for center 2 and center 3
        Out_data_center2 = {}
        for pth_id in Out_PATIENTS_Label_center2.keys():
            if 'radio' in args.data_name:
                radio_parameter_center2_name_path = os.path.join(args.radio_feature_path, args.mapname)
                H5_file = h5py.File(os.path.join(radio_parameter_center2_name_path, str(pth_id) + '.h5'), 'r')
                out_radio_feature_center2 = H5_file['f_values'][:][0]
                H5_file.close()

                key2 = args.mapname + '_' + str(pth_id)
                Out_data_center2[key2] = {}
                Out_data_center2[key2]['radio_feature'] = out_radio_feature_center2
                Out_data_center2[key2]['label'] = Out_PATIENTS_Label_center2[str(pth_id)]

                if 'open-mtm-dsfr' in args.data_name:
                    dsfr_parameter_center2_name_path = os.path.join(args.dsfr_feature_path, args.mapname)
                    H5_file_dsfr = h5py.File(os.path.join(dsfr_parameter_center2_name_path, str(pth_id) + '.h5'), 'r')
                    dsfr_feature_center2 = H5_file_dsfr['f_values'][:]
                    H5_file_dsfr.close()
                    Out_data_center2[key2]['dsfr_feature'] = dsfr_feature_center2

        # Print the feature dimensions
        print('The feature dimension of center 1 (internal) data is {}, and the period is {}'\
              .format(len(In_data), args.mapname))
        print('The feature dimension of center 2 and center 3 (test) data is {}, and the period is {}'\
              .format(len(Out_data_center2), args.mapname))
        # ================================== Data Preparation ==========================================
        if 'open-mtm-dsfr' in args.data_name:
            self.dsfr = True
        else:
            self.dsfr = False

        # center 1
        information_center1, dsfrfeature_center1, radiofeature_center1 = self.dict2np(In_data, dsfr = self.dsfr)
        print('in data_1 num    0:1=={}:{}'.format(int(len(Inner_PATIENTS_Label) - in_p_num), in_p_num))

        # center 2 and center 3
        information_center2, dsfrfeature_center2, radiofeature_center2 = self.dict2np(Out_data_center2, dsfr= self.dsfr)
        print('out data_2 num    0:1=={}:{}'.format(int(len(Out_PATIENTS_Label_center2) - out_p_num_sy), out_p_num_sy))

        # read radiomics name
        df1 = pd.read_excel(args.feature_excel_path)  # Open the feature names excel file
        data_feature_name = list(df1.columns.values) # Convert the feature names to list
        dsfr_data_feature_name = [i for i in range(dsfrfeature_center1[0].shape[1])] # Create a list of feature indices

        # ====================================== Random grid parameter search =============================================
        flag_1 = 0
        excel_row_count = 1
        for iii in range(args.N):
            flag_1 += 1

            print('===Round:', flag_1, '/', args.N, '===')
            gc.collect()

            lasso_alpha = random.uniform(0.0001, 0.1)
            lasso_alpha2 = random.uniform(0.0001, 0.1)

            pd_all_save_before = {}
            feature_save_information = {}

            X_train_indices = np.squeeze(np.argwhere(np.array(in_type) == 1))
            X_val_indices = np.squeeze(np.argwhere(np.array(in_type) == 2))

            train = X_train_indices
            validation = X_val_indices
            excel_row_count = excel_row_count + 1

            phase_filename = args.mapname
            self.radio_savename = []
            self.dsfr_savename = []

            # center 1 all index
            id_center1 = information_center1[:, 1]
            label_center1 = information_center1[:, 2].astype(int)
            radiofeature_center1_single = radiofeature_center1.astype('float64')

            # center1 train
            id_center1_train = id_center1[train]
            label_center1_train = label_center1[train]
            radiofeature_center1_single_train = radiofeature_center1_single[train]
            type_train = np.ones([label_center1_train.shape[0],1])
            data_df_train = np.concatenate((id_center1_train.reshape(-1,1),type_train,label_center1_train.reshape(-1,1),radiofeature_center1_single_train), axis=1)

            # center 1 validation
            id_center1_vali = id_center1[validation]
            label_center1_vali = label_center1[validation]
            radiofeature_center1_single_vali = radiofeature_center1_single[validation]
            type_validation = 2*np.ones([label_center1_vali.shape[0], 1])
            data_df_validation = np.concatenate((id_center1_vali.reshape(-1, 1), type_validation,
                                            label_center1_vali.reshape(-1, 1), radiofeature_center1_single_vali),
                                           axis=1)

            # center 2 3
            index2 = np.argwhere(information_center2 == phase_filename)
            id_center2 = information_center2[index2[:, 0], 1]
            label_center2 = information_center2[index2[:, 0], 2].astype(int)
            radiofeature_center2_single = radiofeature_center2[index2[:, 0], :].astype('float64')
            type_test = 3*np.ones([label_center2.shape[0], 1])
            data_df_test = np.concatenate((id_center2.reshape(-1, 1), type_test,
                                            label_center2.reshape(-1, 1), radiofeature_center2_single),
                                           axis=1)
            feature_infor = np.concatenate((data_df_train, data_df_validation, data_df_test), axis=0)
            # with open(os.path.join(args.savepath, 'feature_information_all.csv'), 'w', newline='') as f:
            #     writer = csv.writer(f)
            #     writer.writerow(['ID', 'Type', 'Label'] + data_feature_name)
            #     writer.writerows(feature_infor)

            if 'open-mtm-dsfr' in args.data_name:
                dsfrfeature_center1_single_train = dsfrfeature_center1.astype('float64')[train]
                dsfrfeature_center1_single_valid = dsfrfeature_center1.astype('float64')[validation]
                dsfrfeature_center2_single = dsfrfeature_center2[index2[:, 0], :].astype('float64')

            # Normalization
            # Normalize external values first, otherwise the maximum and minimum values will change after internal values are overwritten
            radiofeature_out_center2_be = (radiofeature_center2_single - radiofeature_center1_single_train.min(axis=0)) / (
                    radiofeature_center1_single_train.max(axis=0) - radiofeature_center1_single_train.min(axis=0) + 1e-12)
            radiofeature_center1_gy_be_vali = (radiofeature_center1_single_vali - radiofeature_center1_single_train.min(
                axis=0)) / (radiofeature_center1_single_train.max(axis=0) - radiofeature_center1_single_train.min(axis=0) + 1e-12)
            # center 1
            radiofeature_center1_gy_be = (radiofeature_center1_single_train - radiofeature_center1_single_train.min(axis=0)) / (
                        radiofeature_center1_single_train.max(axis=0) - radiofeature_center1_single_train.min(axis=0) + 1e-12)

            if 'open-mtm-dsfr' in args.data_name:
                dsfrfeature_out_center2 = (dsfrfeature_center2_single - dsfrfeature_center1_single_train.min(
                    axis=0)) / (dsfrfeature_center1_single_train.max(axis=0) - dsfrfeature_center1_single_train.min(
                    axis=0) + 1e-12)
                dsfrfeature_center1_gy_vali = (dsfrfeature_center1_single_valid - dsfrfeature_center1_single_train.min(axis=0)) / \
                                         (dsfrfeature_center1_single_train.max(axis=0) - dsfrfeature_center1_single_train.min(axis=0) + 1e-12)
                dsfrfeature_center1_gy = (dsfrfeature_center1_single_train - dsfrfeature_center1_single_train.min(axis=0)) / \
                                         (dsfrfeature_center1_single_train.max(axis=0) - dsfrfeature_center1_single_train.min(axis=0) + 1e-12)

        #  ====================================== ICC select radiomics features  ======================================
            # Excluding ICC < 0.8
            # Intra icc
            radio_icc_1_1 = pd.read_excel(os.path.join(args.icc_file, 'Intra_'+phase_filename + '.xlsx'))
            radio_icc_1_1 = radio_icc_1_1['icc_score'].values
            radio_icc_1_1[radio_icc_1_1 >= 0.8] = 1
            radio_icc_1_1[radio_icc_1_1 < 0.8] = 0
            radio_icc_index_1_1 = [x for x in range(len(radio_icc_1_1)) if
                                   radio_icc_1_1[x] == 1]
            # Inter icc
            radio_icc_1_2 = pd.read_excel(os.path.join(args.icc_file, 'Inter_'+phase_filename + '.xlsx'))
            radio_icc_1_2 = radio_icc_1_2['icc_score'].values
            radio_icc_1_2[radio_icc_1_2 >= 0.8] = 1
            radio_icc_1_2[radio_icc_1_2 < 0.8] = 0
            radio_icc_index_1_2 = [x for x in range(len(radio_icc_1_2)) if
                                   radio_icc_1_2[x] == 1]
            # icc finally
            radio_icc_index = [i for i in radio_icc_index_1_1 if i in radio_icc_index_1_2]

            radiofeature = np.zeros([radiofeature_center1_gy_be.shape[0], len(radio_icc_index)])
            radiofeature_validation = np.zeros([radiofeature_center1_gy_be_vali.shape[0], len(radio_icc_index)])
            radiofeature_out = np.zeros([radiofeature_out_center2_be.shape[0], len(radio_icc_index)])

            for need in range(len(radio_icc_index)):
                radiofeature[:, need] = radiofeature_center1_gy_be[:, radio_icc_index[need]]
                radiofeature_validation[:, need] = radiofeature_center1_gy_be_vali[:, radio_icc_index[need]]
                radiofeature_out[:, need] = radiofeature_out_center2_be[:, radio_icc_index[need]]

            # input features
            radiofeature_center1_gy = radiofeature
            radiofeature_center1_gy_validation = radiofeature_validation
            radiofeature_out_center2 = radiofeature_out
            # ====================================== First features selecting =============================================
            # Preliminary selecting
            inner_radio_feature_choose_center1, inner_radio_feature_choose_center1_valid, out_radio_feature_choose_center2, random_radio, radio_index = \
                self.feature_choose(radiofeature_center1_gy, radiofeature_center1_gy_validation, radiofeature_out_center2,
                                        label_center1_train, pers_chi=40)
            # Radiomics: Indices of features retained after the first screening
            radio_coef_nozero = radio_index
            # radiomics：Indices of original features retained after the first screening, corresponding to the ICC-selected features in the previous step
            first_radio_index = [radio_icc_index[i] for i in radio_coef_nozero]

            if 'open-mtm-dsfr' in args.data_name:
                inner_dsfr_feature_choose_center1, inner_dsfr_feature_choose_center1_validation, out_dsfr_feature_choose_center2, random_dsfr, dsfr_index= \
                    self.feature_choose(np.squeeze(dsfrfeature_center1_gy), np.squeeze(dsfrfeature_center1_gy_vali),np.squeeze(dsfrfeature_out_center2),
                                             label_center1_train, pers_chi=40)\
                # deep learning: Indices of features retained after the first screening
                first_dsfr_index = dsfr_index
            # ====================================== Second features selecting =============================================
            try:
                os.makedirs(os.path.join(args.savepath, 'radio_lasso' + phase_filename), exist_ok=True)
                first_mser, inner_radio_feature_finally_center1, inner_radio_feature_finally_center1_validation, out_radio_feature_finally_center2, fl_radio, radio_coef, radio_intercept = \
                    self.choose_feature(inner_radio_feature_choose_center1, label_center1_train, \
                                        feature_out = out_radio_feature_choose_center2, \
                                        feature_valid = inner_radio_feature_choose_center1_valid,
                                        alpha=lasso_alpha)
                # If the number of retained features is 0, skip this round.
                if fl_radio==0:
                    continue
                print(phase_filename + ' the nunmber of radiomics features after the second feature selection:' + str(inner_radio_feature_finally_center1.shape[1]))
                radio_coef_nozero = np.nonzero(radio_coef)[0] # feature index
                second_radio_index = [first_radio_index[i] for i in radio_coef_nozero]

                # Saving the total rad-score for the R code classifier
                pd_all_save_before[phase_filename] = {}
                # Saving the total radiomics features
                pd_all_save_before[phase_filename]['radio_feature_center1'] = inner_radio_feature_finally_center1
                pd_all_save_before[phase_filename]['radio_feature_center1_valid'] = inner_radio_feature_finally_center1_validation
                pd_all_save_before[phase_filename]['radio_feature_center2'] = out_radio_feature_finally_center2

                if 'open-mtm-dsfr' in args.data_name:
                    os.makedirs(os.path.join(args.savepath, 'dsfr_lasso' + phase_filename), exist_ok=True)
                    first_msed, inner_dsfr_feature_finally_center1, inner_dsfr_feature_finally_center1_validation,out_dsfr_feature_finally_center2, fl_dsfr, dsfr_coef, dsfr_intercept = \
                        self.choose_feature(inner_dsfr_feature_choose_center1, label_center1_train, \
                                            feature_out=out_dsfr_feature_choose_center2, \
                                            feature_valid=inner_dsfr_feature_choose_center1_validation,
                                            alpha=lasso_alpha)
                    if fl_dsfr == 0:
                        continue
                    print(phase_filename + ' the nunmber of deep learning features after the second feature selection: ' + str(inner_dsfr_feature_finally_center1.shape[1]))
                    dsfr_coef_nozero = np.nonzero(dsfr_coef)[0]
                    second_dsfr_index = [first_dsfr_index[h] for h in dsfr_coef_nozero]

                    # Saving the total deep learning features
                    pd_all_save_before[phase_filename]['deeplearning_feature_center1'] = np.array(inner_dsfr_feature_finally_center1)
                    pd_all_save_before[phase_filename]['deeplearning_feature_center1_valid'] = np.array(inner_dsfr_feature_finally_center1_validation)
                    pd_all_save_before[phase_filename]['deeplearning_feature_center2'] = np.array(out_dsfr_feature_finally_center2)

                print(phase_filename + ' the nunmber of radiomics and deep leanring features after the second feature selection: ' + str(inner_radio_feature_finally_center1.shape[1]+inner_dsfr_feature_finally_center1.shape[1]))

                # Saving the index after selecting
                feature_save_information[phase_filename + '_radio_index'] = second_radio_index
                feature_save_information[phase_filename + '_deeplearning_index'] = second_dsfr_index

                rad_signature_index = feature_save_information[args.mapname + '_radio_index']
                rad_feature_name = [data_feature_name[h] for h in rad_signature_index]
                if 'open-mtm-dsfr' in args.data_name:
                    deeplearning_signature_index = list(feature_save_information[args.mapname + '_deeplearning_index'])
                    # Concatenating the radiomics and deep learning features
                    center1_signature = np.hstack([pd_all_save_before[args.mapname]['radio_feature_center1'], pd_all_save_before[args.mapname]['deeplearning_feature_center1']])
                    center1_signature_valid = np.hstack([pd_all_save_before[args.mapname]['radio_feature_center1_valid'], pd_all_save_before[args.mapname]['deeplearning_feature_center1_valid']])
                    center2_signature = np.hstack([pd_all_save_before[args.mapname]['radio_feature_center2'], pd_all_save_before[args.mapname]['deeplearning_feature_center2']])
                    # Concatenating the radiomics and deep learning feature's name
                    signature_name = rad_feature_name + deeplearning_signature_index
                else:
                    center1_signature = pd_all_save_before[args.mapname]['radio_feature_center1']
                    center1_signature_valid = pd_all_save_before[args.mapname]['radio_feature_center1_valid']
                    center2_signature = pd_all_save_before[args.mapname]['radio_feature_center2']
                    signature_name = rad_feature_name

                # Selecting the features by lasso finally (radiomics and deep learning together)
                second_mse, pred_feature_center1_lasso, pred_feature_center1_lasso_valid, pred_feature_center2_lasso, fl_finally, finally_coef, finally_intercept = \
                    self.choose_feature(center1_signature, label_center1_train, \
                                        feature_out=center2_signature, \
                                        feature_valid = center1_signature_valid,\
                                        alpha=lasso_alpha2)
                if fl_finally == 0:
                    continue
                # the number of features finally
                print(args.mapname + ' the nunmber of radiomics and deep leanring features after feature selection finally: ' + str(pred_feature_center1_lasso.shape[1]))

                # # saving the number of features finally
                # final_feature_dimension = pred_feature_center1_lasso.shape[1]

                finally_coef_nozero = np.nonzero(finally_coef)[0]  #  saving the index of features finally
                finally_signature_name = [signature_name[h] for h in finally_coef_nozero]

                # Calculating the rad-score in center1 train
                rad_center1_list = []
                for hh in range(center1_signature.shape[0]):
                    weight_feature_center1 = 0.0
                    for mm in range(center1_signature.shape[1]):
                        weight_feature_center1 = weight_feature_center1 + \
                                                         center1_signature[hh][mm] * \
                                                         finally_coef[mm]
                    rad_center1 = finally_intercept + weight_feature_center1
                    rad_center1_list.append(rad_center1)  # radscore

                # Calculating the rad-score in center1 validation
                rad_center1_list_valid = []
                for kk in range(center1_signature_valid.shape[0]):
                    weight_feature_center1_valid = 0.0
                    for nn in range(center1_signature_valid.shape[1]):
                        weight_feature_center1_valid = weight_feature_center1_valid + \
                                                         center1_signature_valid[kk][nn] * \
                                                         finally_coef[nn]
                    rad_center1_valid = finally_intercept + weight_feature_center1_valid
                    rad_center1_list_valid.append(rad_center1_valid)  # radscore

                # Calculating the rad-score in center2 and center3
                rad_center2_list = []
                for hh in range(center2_signature.shape[0]):
                    weight_feature_center2 = 0.0
                    for mm in range(center2_signature.shape[1]):
                        weight_feature_center2 = weight_feature_center2 + \
                                                         center2_signature[hh][mm] * \
                                                         finally_coef[mm]
                    rad_center2 = finally_intercept + weight_feature_center2
                    rad_center2_list.append(rad_center2)  # radscore

                id_center_all = np.concatenate([id_center1_train, id_center1_vali, id_center2], axis=0)
                rad_score_all = np.array(rad_center1_list + rad_center1_list_valid + rad_center2_list)
                merged_arr = np.column_stack((id_center_all, rad_score_all))
                os.makedirs(os.path.join(args.savepath, 'Radscore'), exist_ok=True)

                # save radscore information
                with open(os.path.join(args.savepath, 'Radscore', 'round(%d)_mse(%f).csv' % (flag_1, second_mse)), 'w', newline='') as f:
                    writer = csv.writer(f)
                    writer.writerow(['ID', 'radscore'])
                    writer.writerows(merged_arr)

                # # save mse, lasso_alpha, lasso_alpha2 and others
                # results_savepath = os.path.join(args.savepath, 'results.xlsx')
                # if flag_1 == 1:
                #     self.writesimple(results_savepath, 'round', 1, 1)
                #     self.writesimple(results_savepath, 'mse', 1, 2)
                #     self.writesimple(results_savepath, 'first_lasso_alpha', 1, 3)
                #     self.writesimple(results_savepath, 'second_lasso_alpha', 1, 4)
                # self.writesimple(results_savepath, str(flag_1), flag_1+1, 1)
                # self.writesimple(results_savepath, str(second_mse), flag_1+1, 2)
                # self.writesimple(results_savepath, str(lasso_alpha), flag_1+1, 3)
                # self.writesimple(results_savepath, str(lasso_alpha2), flag_1+1, 4)
            except:
                print('No features!')

if __name__=="__main__":
    root_path = r'./DSFR_MTM_Radiology'
    key_word = "ID"
    map_name = "PVP_ID"
    parser = argparse.ArgumentParser(description="MTM_classification")
    parser.add_argument('--mapname', type=str, default=map_name)
    parser.add_argument('--split_train_test', type=str, metavar='PATH',
                        default=os.path.join(root_path, 'datainfo.csv'))
    parser.add_argument('--icc_file', type=str, metavar='PATH',
                        default=os.path.join(root_path, 'ICC'))
    parser.add_argument('--savepath', type=str, metavar='PATH',
                        default=os.path.join(root_path, 'results', 'feature_select', map_name),
                        help='results save path')
    parser.add_argument('--radio_feature_path', type=str, metavar='PATH',
                        default=os.path.join(root_path, 'Radiomic', key_word),
                        help='radiomics feature load path')
    parser.add_argument('--dsfr_feature_path', type=str, metavar='PATH',
                        default=os.path.join(root_path, 'UCtransNet', key_word),
                        help='deep learning feature load path')
    parser.add_argument('--feature_excel_path', type=str, metavar='PATH',
                        default=os.path.join(root_path, 'feature_title.xlsx'), help='radiomics feature name load path')
    parser.add_argument('--data_name', type=list,
                        default=['radio', 'open-mtm-dsfr'])
    parser.add_argument('--N', type=int,
                        default=5000, help="random grid times")
    Classfier().run(parser.parse_args())