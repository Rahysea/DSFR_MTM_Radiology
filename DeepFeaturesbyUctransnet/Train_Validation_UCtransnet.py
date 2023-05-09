# -*- coding: utf-8 -*-
# @Time    : 2022/5/19 19:30
# @Author  : Atsea
# @File    : Train_Validation_uctransnet.py
# @Software: PyCharm

from loss_dice.dsc import *
from loss_dice.Dice_loss import WeightedDiceBCE
from model.UCTransNet import *
from model.Config import get_CTranS_config
from dataloader.dataloader_seg import data_loaders
from openpyxl import Workbook
from openpyxl import load_workbook
import os
import time
import argparse
os.environ["CUDA_VISIBLE_DEVICES"] = "1"
os.environ['PYTHONHASHSEED'] = str(666)

def log_loss_summary(loss, step, prefix=""):
    print("epoch {} | {}: {}".format(step + 1, prefix + "loss", np.mean(loss)))

def writesimple(path, data, index, column, sheetname='Sheet'):
    index = index +1
    column = column+1
    if os.path.exists(path):
        bg = load_workbook(path)
        sheets = bg.sheetnames
        if sheetname in sheets:
            sheet = bg[sheetname]
            sheet.cell(index, column, data)
            bg.save(path)
            bg.close()
        else:
            sheet = bg.create_sheet(sheetname)
            sheet.cell(index, column, data)
            bg.save(path)
            bg.close()
    else:
        bg = Workbook()
        bg1 = bg['Sheet']
        bg1.title = sheetname
        sheet = bg[sheetname]
        sheet.cell(index, column, data)
        bg.save(path)
        bg.close()

def main():
    key_word = 'AP_M40'

    datapath = r'./AP_M40'
    csvpath = os.path.join(os.path.abspath('..'), 'datainfo.csv')
    resultsave = os.path.join(os.path.abspath('..'), 'Weight_model', key_word)

    parser = argparse.ArgumentParser()
    parser.add_argument('--root', default=datapath, help='path to dataset (images list file)')
    parser.add_argument('--csv',default=csvpath, help='path to csv (devide train and test datasets)')
    parser.add_argument('--lr', default=0.0001, type=float, help='learning rate for training')
    parser.add_argument('--epochs', default=200, type=int, help='number of total epochs to run')
    parser.add_argument('--start_epoch', default=0, type=int, help='manual epoch number (useful on restarts)')
    parser.add_argument('--batch_size', type=int, default=5, help='input batch size')
    parser.add_argument('--optim', type=str, default='Adam', help='optim for training, Adam / SGD (default)')
    parser.add_argument('--momentum', default=0.9, type=float, help='momentum for SGD')
    parser.add_argument('--weight_decay', default=1e-4, type=float, help='weight_decay for SGD / Adam')
    parser.add_argument('--gpu', type=bool, default=True, help='use GPU or not')
    args = parser.parse_args()

    datapath = args.root
    csvpath = args.csv

    os.makedirs(resultsave, exist_ok=True)
    excelsavepath = os.path.join(resultsave, 'results.xlsx')

    config_vit = get_CTranS_config()
    config_vit.batch_size = args.batch_size
    config_vit.epochs = args.epochs
    ucnet = UCTransNet(config=config_vit, n_channels=config_vit.n_channels, n_classes=config_vit.n_labels, img_size=config_vit.img_size)
    if args.gpu:
        device = ('cuda' if torch.cuda.is_available() else 'cpu')
        ucnet.to(device)

    # 0.5*BCE and 0.5*DICE loss
    criterion = WeightedDiceBCE(dice_weight=0.5,BCE_weight=0.5)
    best_validation_dsc = 0.0

    # optim
    if args.optim == 'SGD':
        optimizer = torch.optim.SGD(ucnet.parameters(), lr=args.lr, momentum=args.momentum, weight_decay=args.weight_decay)
    elif args.optim == 'Adam':
        optimizer = torch.optim.Adam(ucnet.parameters(), lr=args.lr, weight_decay=args.weight_decay)
    else:
        raise NotImplementedError('Other optimizer is not implemented')

    # excel item save
    writesimple(excelsavepath, 'epoch', 0, 0, 'train')
    writesimple(excelsavepath, 'loss', 0, 1, 'train')
    writesimple(excelsavepath, 'epoch', 0, 0, 'test')
    writesimple(excelsavepath, 'loss', 0, 1, 'test')
    writesimple(excelsavepath, 'mean_dsc', 0, 2, 'test')

    # datasets
    step = 0
    loss_train = []
    loss_valid = []
    dataloader_train, dataloader_valid = data_loaders(datapath, csvpath, task='MTM', batch_size=args.batch_size, workers=20)
    loaders = {'train': dataloader_train, 'valid': dataloader_valid}
    for epoch in range(args.start_epoch, args.epochs):
        a =time.time()
        for phase in ["train", "valid"]:
            if phase == "train":
                ucnet.train()
            else:
                ucnet.eval()
            valid_pred = []
            valid_true = []
            # get data
            for i, item in enumerate(loaders[phase]):
                if phase == "train":
                    step += 1
                data, label = item
                if torch.cuda.is_available():
                    data = data.to(device)
                    label = label.to(device)
                    # Clear the gradient values
                    optimizer.zero_grad()
                    # Backpropagation is only used during training.
                    with torch.set_grad_enabled(phase == "train"):
                        pred = ucnet(data)
                        loss = criterion(pred, label)
                        if phase == 'valid':
                            loss_valid.append(loss.item())
                            y_pred_np = pred.detach().cpu().numpy()
                            valid_pred.extend([y_pred_np[s] for s in range(y_pred_np.shape[0])])
                            y_true_np = label.detach().cpu().numpy()
                            valid_true.extend([y_true_np[s] for s in range(y_true_np.shape[0])])
                        elif phase == 'train':
                            loss_train.append(loss.item())
                            loss.backward() # Compute the parameter gradients via backpropagation.
                            optimizer.step() # Update the parameters via gradient descent.
                print('epoch:{}/{}, type:{}, step:{}, loss:{}'.format(epoch+1, args.epochs-args.start_epoch, phase, step, loss.item()))

            if phase == "train":
                log_loss_summary(loss_train, epoch)
                writesimple(excelsavepath, epoch, epoch+1, 0, 'train')
                writesimple(excelsavepath, np.mean(loss_train), epoch+1, 1, 'train')
                loss_train = []
            if phase == "valid":
                log_loss_summary(loss_valid, epoch, prefix="val_")
                # calculate mean dice
                mean_dsc = np.mean(
                    dsc_per_volume(
                        valid_pred,
                        valid_true,
                        dataloader_valid.dataset.chooseslicepath
                    )
                )
                log_scalar_summary("val_dsc", mean_dsc, epoch)
                if mean_dsc > best_validation_dsc:
                    best_validation_dsc = mean_dsc
                    torch.save(ucnet.state_dict(), os.path.join(resultsave, "uctransnet.pt"))
                writesimple(excelsavepath, epoch, epoch+1, 0, 'test')
                writesimple(excelsavepath, np.mean(loss_valid), epoch+1, 1, 'test')
                writesimple(excelsavepath, np.mean(mean_dsc), epoch+1, 2, 'test')
                loss_valid = []
                # train results
        print("time: ",time.time() - a,' s')

    print("\nBest validation mean DSC: {:4f}\n".format(best_validation_dsc))

if __name__ == '__main__':
    main()